from this import d
import numpy as np
import random
from math import sqrt
import cmath
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from google.colab import files


distances = []
newDistances = []
refPoints = []
coordsUnknownPoint = []
# distances = [3.857, 3.988, 3.497] # Replica paper
# refPoints = [[27.297, -4.953, 1.47], [20.693, -4.849, 1.93], [22.59, 0.524, 1.2], [17.113, -3.003, 2.17], [22.554, 4.727, 1.77]] # Replica paper
# coordsUnknownPoint = [26.7726, -1.3389] # Replica paper
FinalSolution = []
FinalSolution1 = []
FinalSolution2 = []
t = any
t1 = any
t2 = any
A_segundoGrado = 0
B_segundoGrado = 0
C_segundoGrado = 0
x_max = 0
x_min = 0
y_max = 0
y_min = 0
z_max = 0
z_min = 0
w_max = 0
w_min = 0
error = 0
ruta_archivo = "coordenadas.xlsx"
# nombres_variables = ["Known coords", "Ancla1", "Ancla2", "Ancla3", "Calculated coords", "Error"]
nombres_variables = ["X", "Y", "Z", "W", "Distances"]
toInsert = []

# coordsUnknownPoint = [7, 10]; #Conozco solo 2
# coordsUnknownPoint = [7, 10, 2.1, -1.4]; #Conozco las 4

def generar_excel(coords, nombres_variables, ruta_archivo):
    workbook = Workbook()  # Crea un objeto Workbook

    # Accede a la hoja de cálculo activa
    worksheet = workbook.active
    worksheet.title = "Coordenadas"  # Cambia el nombre de la hoja de cálculo
    
    # Escribe los nombres de las variables en la primera fila
    for col, nombre_variable in enumerate(nombres_variables):
        worksheet.cell(row=1, column=col + 2, value=nombre_variable)

    # Escribe las coordenadas en las columnas correspondientes
    for row, coord in enumerate(coords):
        if row == 0:
            worksheet.cell(row=row + 2, column=1, value="Known Coords")
        elif row == 1:
            worksheet.cell(row=row + 2, column=1, value="Calculated Coords")
        elif row == 2:
            worksheet.cell(row=row + 2, column=1, value="Measured Error")
        else:
            worksheet.cell(row=row + 2, column=1, value="Ancla " + str(row - 2))
        for col, valor in enumerate(coord):
            worksheet.cell(row=row + 2, column=col + 2, value=valor)

    workbook.save(ruta_archivo)
        # Descarga el archivo Excel generado
    files.download(ruta_archivo)



# Generar puntos ancla aleatorios
def generadorPuntosAnclaRandom(coords):
    subarray = [random.uniform(-10, 10) for _ in range(coords)]
    return subarray

refPoints = [generadorPuntosAnclaRandom(4) for _ in range(3)]
coordsUnknownPoint = generadorPuntosAnclaRandom(4)
print("Coordenadas conocidas: ", coordsUnknownPoint)

# Matriz Identidad
I = np.identity(len(refPoints[0]))

# Matrices para el sist. de ecuaciones
A = np.array([])
b = np.array([])

# Función que calcule las distancias
def calculateDistances():
  global distances
  for coords in refPoints:
    value = 0
    for i in range(len(coordsUnknownPoint)):
      value += (coords[i]-coordsUnknownPoint[i])**2
    s = sqrt(value)
    distances.append(s)
      
  print("Distancias: ", distances)

def calcular_error(coordenadas_conocidas, coordenadas_calculadas):
    if len(coordenadas_conocidas) != len(coordenadas_calculadas):
        raise ValueError("Los arrays de coordenadas deben tener la misma longitud.")

    errores = np.abs(coordenadas_conocidas - coordenadas_calculadas)
    return errores

def calculateNewDistances():
  global newDistances
  for coords in refPoints:
    value = 0
    if len(FinalSolution1) > 0:
      for i in range(len(FinalSolution1)):
        value += (coords[i]-FinalSolution1[i])**2
      s = sqrt(value)
      newDistances.append(s)
    else:
      for i in range(len(FinalSolution)):
        value += (coords[i]-FinalSolution[i])**2
      s = sqrt(value)
      newDistances.append(s)
      
  print("Las primeras distancias calculadas: ",distances)
  print("Las nuevas distancias calculadas: ",newDistances)

# Generar las ecuaciones de A en función de los puntos de referencia
def generateAEcs():
  global A
  for coords in refPoints:
    newRow = [1]
    for value in coords:
      newRow.append(-2*value)
    print(newRow)
    A = np.append(A, np.array(newRow))
    A = np.reshape(A, (-1, len(newRow)))
  # print("A =", A)

# Generar las ecuaciones de B en función de los puntos de referencia y distancias
def generateBEcs():
  global b
  for coords, dist in zip(refPoints, distances):
    newRow = dist**2
    for value in coords:
      newRow = newRow - value**2
    b = np.append(b, newRow)
  # print("b =", b)

calculateDistances()
generateAEcs()
generateBEcs()

# Calculate EC 2º Grado values
def secondGradeVariables(x_p, x_h):
  global A_segundoGrado, B_segundoGrado, C_segundoGrado

  if len(x_p) != len(x_h):
    raise ValueError("La longitud debe ser la misma")
  else:
    length = len(x_p)

  A_segundoGrado = sum([x_h[i]**2 for i in range(1,length)])
  # print(A_segundoGrado)

  B_segundoGrado = sum([2*x_p[i]*x_h[i] for i in range(1,length)]) - x_h[0]
  # print(B_segundoGrado)

  C_segundoGrado = sum([x_p[i]**2 for i in range(1,length)]) - x_p[0]
  # print(C_segundoGrado)


# Ec 2º Grado
def solve_quadratic(a, b):
  global t, t1, t2

  if insideSqrt < 0:
    solucion = (-b)/(2*a)
    return [solucion]
  else:
    solucion1 = (-b-cmath.sqrt(insideSqrt))/(2*a)
    solucion2 = (-b+cmath.sqrt(insideSqrt))/(2*a)
    return [solucion1.real, solucion2.real]

def calculateDifferences(x1, x2):
  global d1, d2

  d1 = x1[0] - sum([x1[i]**2 for i in range(1,len(x1))])
  d2 = x2[0] - sum([x2[i]**2 for i in range(1,len(x2))])

  # print("dif1 =", d1)
  # print("dif2 =", d2)

def findSolution():
  X = np.array(N1)
  Y = np.array(N2)
  P = np.array(refPoints)

  res1 = 0
  res2 = 0
  res = []

  for i in range(len(P)):
    res1 += (np.linalg.norm(X - P[i])**2 - distances[i])**2
    res2 += (np.linalg.norm(Y - P[i])**2 - distances[i])**2

  res.append(res1)
  res.append(res2)

  indexMinValue = np.argmin(res)

  # print("minimice =", res[indexMinValue])

  if indexMinValue == 0:
    print("La solución es N1")
  else:
    print("La solución es N2")


# Fin de la definición de funciones, aquí empieza el cálculo del punto desconocido

# x_p (soluciones particulares) y x_h (soluciones homogéneas) usando la pseudo-inversa
A_pseudo_inv = np.linalg.pinv(A)
print(A_pseudo_inv)
x_p = A_pseudo_inv.dot(b)

# SVD (Descomposición en valores singulares). U matriz unitaria. S matriz diagonal. VT matriz transpuesta conjugada.
U, S, VT = np.linalg.svd(A)
# print("VT = ",VT)
x_h = VT[-1] # x_h último vector en la matriz VT (matriz de vectores propios transpuestos)

# print("x_p =", x_p)
# print("x_h =", x_h)

# Ecuación de 2º Grado
secondGradeVariables(x_p, x_h)
insideSqrt = (B_segundoGrado**2)-(4*A_segundoGrado*C_segundoGrado)
# print("insideSqrt =", insideSqrt)

ecSecondGradeSolution = solve_quadratic(A_segundoGrado, B_segundoGrado)
toInsert = [coordsUnknownPoint]

# toInsert.append(distances)

if len(ecSecondGradeSolution) == 1:
  sol = x_p + ecSecondGradeSolution*x_h
  sol = sol[1:]
  FinalSolution = sol.dot(I)
  print("SOLUCION FINAL =", FinalSolution)
  toInsert.append(FinalSolution)
  error = calcular_error(coordsUnknownPoint, FinalSolution)
else:
  sol1 = x_p + ecSecondGradeSolution[0]*x_h
  sol2 = x_p + ecSecondGradeSolution[1]*x_h
  calculateDifferences(sol1, sol2)
  sol1_copy = sol1
  sol2_copy = sol2
  sol1 = sol1[1:] # Me interesa quedarme con las coordenadas desconocidas, ya que las primeras lo son.
  sol2 = sol2[1:]
  # print(sol1)
  # print(sol2)
  FinalSolution1 = sol1.dot(I)
  FinalSolution2 = sol2.dot(I)
  print("SOLUCION FINAL 1 =", FinalSolution1)
  print("SOLUCION FINAL 2 =", FinalSolution2)
  toInsert.append(FinalSolution1)
  error = calcular_error(coordsUnknownPoint, FinalSolution1)



print("El error cometido es: ", error)

print("Las coordenadas conocidas son: ", coordsUnknownPoint)


toInsert.append(error)
toInsert.extend(refPoints)
print("TO INSERT: ", toInsert)
generar_excel(toInsert, nombres_variables, ruta_archivo)
calculateNewDistances()

# findSolution()

# Representación gráfica

# Calcular los límites del gráfico

def generatePlots(refPoints):
  global x_max, x_min, y_max, y_min, z_max, z_min, w_max, w_min

  refPointsXY = []
  refPointsZW = []
  x_values = []
  y_values = []
  z_values = []
  w_values = []

  for subarray in refPoints:
      refPointsXY.append(subarray[:2])
      refPointsZW.append(subarray[:2])

  print(refPointsZW)
  print(newDistances)

  for i, (x, y) in enumerate(refPointsXY):
        x_values.append(refPointsXY[i][0] + distances[i])
        x_values.append(refPointsXY[i][0] - distances[i])
        y_values.append(refPointsXY[i][1] + distances[i])
        y_values.append(refPointsXY[i][1] - distances[i])
        circle = plt.Circle((x, y), distances[i], color='b', fill=False)
        ax1.add_patch(circle)

  for i, (z, w) in enumerate(refPointsZW):
      z_values.append(refPointsZW[i][0] + newDistances[i])
      z_values.append(refPointsZW[i][0] - newDistances[i])
      w_values.append(refPointsZW[i][1] + newDistances[i])
      w_values.append(refPointsZW[i][1] - newDistances[i])
      circle = plt.Circle((z, w), newDistances[i], color='b', fill=False)
      ax2.add_patch(circle)

  x_max = max(x_values)
  x_min = min(x_values)
  y_max = max(y_values)
  y_min = min(y_values)

  z_max = max(z_values)
  z_min = min(z_values)
  w_max = max(w_values)
  w_min = min(w_values)

# Crear los gráficos y los subplots
fig, (ax1, ax2) = plt.subplots(1, 2)

# Añadir las circunferencias al primer gráfico
generatePlots(refPoints)
if len(FinalSolution2) > 0 :
  ax1.plot(FinalSolution1[0], FinalSolution1[1], 'ro')
  ax1.plot(FinalSolution2[0], FinalSolution2[1], 'yo')
else:
  ax1.plot(FinalSolution[0], FinalSolution[1], 'ro')
ax1.plot(coordsUnknownPoint[0], coordsUnknownPoint[1], 'go')
ax1.set_xlabel('x')
ax1.set_ylabel('y')

# Configurar los límites de los ejes del primer gráfico
ax1.set_xlim(x_min, x_max)
ax1.set_ylim(y_min, y_max)
ax1.axis('equal')

# Añadir las circunferencias al segundo gráfico
if len(FinalSolution2) > 0 :
  ax2.plot(FinalSolution1[0], FinalSolution1[1], 'ro')
  ax2.plot(FinalSolution2[0], FinalSolution2[1], 'yo')
else:
  ax2.plot(FinalSolution[0], FinalSolution[1], 'ro')
# ax2.plot(coordsUnknownPoint[2], coordsUnknownPoint[3], 'go')
ax2.plot(coordsUnknownPoint[0], coordsUnknownPoint[1], 'go')
ax2.set_xlabel('z')
ax2.set_ylabel('w')

# Configurar los límites de los ejes del segundo gráfico
ax2.set_xlim(z_min, z_max)
ax2.set_ylim(w_min, w_max)
ax2.axis('equal')

# Mostrar los gráficos
plt.show()
generar_excel(toInsert, nombres_variables, ruta_archivo)