## CÓDIGO QUE REPLICA EL PAPER
## SE PASARAN LOS MISMOS DATOS

import os
import numpy as np
import random
import cmath
import time
import pandas as pd
import matplotlib.pyplot as plt
from this import d
from math import sqrt
from openpyxl import Workbook
from google.colab import files

# Función para redirigir la salida de print a un archivo
def redirect_print_to_file(file_name):
    import sys
    f = open(file_name, 'w')
    sys.stdout = f
    return f

log_file = redirect_print_to_file('output_log.txt')
radioCircunferenciaDistancias = 1
distanciaPuntoDesconocido = 1
acumulatedDistanceError = []
performance_multilateralizacion = []
numero_de_iteraciones = 1
A_segundoGrado = 0
B_segundoGrado = 0
C_segundoGrado = 0

A = np.array([])
b = np.array([])

numero_coordenadas = 4
distancias = []
distancias_MEDIDAS_a_cada_punto_ancla = []
distancias_REALES_a_cada_punto_ancla= []
mean2Coords = [0, 0]
mean = [0, 0, 0, 0]
covarianceMatrix2Coords = [[1,0.9], [0.9,1]]
covHighDependency = [
        [1, 0.95, 0.93, 0.92],   # Varianza de x y covarianza entre x y otras dimensiones
        [0.95, 1, 0.94, 0.91],   # Varianza de y y covarianza entre y y z
        [0.93, 0.94, 1, 0.96],   # Varianza de z y covarianza entre z y y
        [0.92, 0.91, 0.96, 1]    # Varianza de w y covarianza entre w y y, y w y z
]

covDependency = [
        [1, 1, 1, 1],   # Varianza de x y covarianza entre x y otras dimensiones
        [1, 1, 1, 1],   # Varianza de y y covarianza entre y y z
        [1, 1, 1, 1],   # Varianza de z y covarianza entre z y y
        [1, 1, 1, 1]    # Varianza de w y covarianza entre w y y, y w y z
]

## Generar coordenadas conocidas (2) del punto desconocido con dependencia entre ellas
###  Necesario?
def generadorCoordenadasConocidasPuntoDesconocido():
    n_samples = 1
    coordenadas_conocidas_punto_desconocido = np.random.multivariate_normal(mean2Coords, covarianceMatrix2Coords, n_samples)
    return coordenadas_conocidas_punto_desconocido

# Generar puntos aleatorios
###  Necesario?
def generadorPuntosRandomUsandoMatrizCovarianzas(mean, covHighDependency, n_samples):
    n_samples = numero_puntos_ancla
    puntos_ancla_generados = np.random.multivariate_normal(mean, covHighDependency, n_samples)
    return puntos_ancla_generados

def generadorPuntosRandomConUnaDistanciaDeterminada(punto_referencia, distancia, numeroDePuntos):
    puntos = []
    for _ in range(numeroDePuntos):
        vector = np.random.rand(4)     
        norm = np.linalg.norm(vector)
        
        # Escalar el vector para que su norma sea igual a la distancia deseada
        vector_escalado = vector / norm * distancia
        
        # Generar el nuevo punto sumando el vector escalado al punto base
        puntoGenerado = punto_referencia + vector_escalado
        
        puntos.append(puntoGenerado)
    
    return np.array(puntos)

def generar_puntos_en_R2(numeroDePuntos, punto_referencia, d1, d2):
    puntosAncla = []

    for _ in range(numeroDePuntos):
        # Distribuir puntos generados
        theta1 = np.random.uniform(0, 2*np.pi)
        theta2 = np.random.uniform(0, 2*np.pi)
        
        x = punto_referencia[0] + d1 * np.cos(theta1)
        y = punto_referencia[1] + d1 * np.sin(theta1)
        z = punto_referencia[2] + d2 * np.cos(theta2)
        w = punto_referencia[3] + d2 * np.sin(theta2)
        
        nuevo_punto = np.array([x, y, z, w])
        puntosAncla.append(nuevo_punto)
        
        # Distancias en los subespacios de R2
        distancia_1 = np.linalg.norm(nuevo_punto[:2] - punto_referencia[:2])
        distancia_2 = np.linalg.norm(nuevo_punto[2:] - punto_referencia[2:])
        
    return np.array(puntosAncla)

# Función que calcula las distancias
def calculateDistances():
  global distancias, coordenadas_punto_desconocido, coordenadas_puntos_ancla_conocidos, distancias_REALES_a_cada_punto_ancla, distancias_MEDIDAS_a_cada_punto_ancla, radioCircunferenciaDistancias
  if len(distancias) > 0:
    distancias = []

  for coords in coordenadas_puntos_ancla_conocidos:
      value = 0
      for i in range(len(coordenadas_punto_desconocido)):
          value += (coords[i]-coordenadas_punto_desconocido[i])**2
      s = sqrt(value)
      distancias.append(s)

  distancias_REALES_a_cada_punto_ancla = np.round(distancias, 3)
  distancias_MEDIDAS_a_cada_punto_ancla = distancias_REALES_a_cada_punto_ancla
  print()
  print("Distancias reales: " + str(distancias_REALES_a_cada_punto_ancla))
  print()

# Calcular la distancia euclidea en el subespacio 2D (considerando solo las coordenadas x e y) ### PROVISIONAL, NO ES DEL TODO CORRECTO, APROXIMACION
def calcular_distancias_subespacio_2D(puntos_conocidos, punto_desconocido):
  global distancias, distancias_REALES_a_cada_punto_ancla, distancias_MEDIDAS_a_cada_punto_ancla, radioCircunferenciaDistancias
  punto_desconocido = np.array(punto_desconocido).reshape(1, -1)

  distanciasSubespacio2D = np.linalg.norm(puntos_conocidos[:, :2] - punto_desconocido[:, :2], axis=1)
  distancias = np.linalg.norm(puntos_conocidos - punto_desconocido, axis=1)
  distancias_REALES_a_cada_punto_ancla = np.round(distancias, 4)
  distancias_MEDIDAS_a_cada_punto_ancla = np.round(distanciasSubespacio2D, 5)

  distanciasIguales = all(x == distancias_MEDIDAS_a_cada_punto_ancla[0] for x in distancias_MEDIDAS_a_cada_punto_ancla)
  if distanciasIguales:
    radioCircunferenciaDistancias = distancias_MEDIDAS_a_cada_punto_ancla[0]


  print("Distancias reales: " + str(distancias_REALES_a_cada_punto_ancla))
  print("Distancias medidas subespacio 2D (considerando solo las coordenadas x e y): " + str(distancias_MEDIDAS_a_cada_punto_ancla))
  print()



# Generar las ecuaciones de A en función de los puntos de referencia
def generateAEcs():
  global A, refPoints
  A = np.array([])
  for coords in refPoints:
    newRow = [1]
    for value in coords:
      newRow.append(-2*value)
    A = np.append(A, np.array(newRow))
    A = np.reshape(A, (-1, len(newRow)))

# Generar las ecuaciones de B en función de los puntos de referencia y distancias
def generateBEcs(distances):
  global b, refPoints
  b = np.array([])
  for coords, dist in zip(refPoints, distances):
    newRow = dist**2
    for value in coords:
      newRow = newRow - value**2
    b = np.append(b, newRow)

# Calcular valores EC 2º Grado
def secondGradeVariables(x_p, x_h):
  global A_segundoGrado, B_segundoGrado, C_segundoGrado

  if A_segundoGrado != 0 and B_segundoGrado != 0 and C_segundoGrado != 0:
    A_segundoGrado = 0
    B_segundoGrado = 0
    C_segundoGrado = 0

  if len(x_p) != len(x_h):
    raise ValueError("La longitud debe ser la misma")
  else:
    length = len(x_p)

  A_segundoGrado = sum([x_h[i]**2 for i in range(1,length)])

  B_segundoGrado = sum([2*x_p[i]*x_h[i] for i in range(1,length)]) - x_h[0]

  C_segundoGrado = sum([x_p[i]**2 for i in range(1,length)]) - x_p[0]

# Ec 2º Grado
def solve_quadratic(a, b, insideSqrt):
  global t, t1, t2

  if insideSqrt < 0:
    solucion = (-b)/(2*a)
    return [solucion]
  else:
    solucion1 = (-b-cmath.sqrt(insideSqrt))/(2*a)
    solucion2 = (-b+cmath.sqrt(insideSqrt))/(2*a)
    return [solucion1.real, solucion2.real]

## En caso de obtener 2 candidatos, esta función determina la solución óptima en base al menor error cometido
def findSolution(FinalSolution1, FinalSolution2):
  global error
  error1 = calcular_error(coordenadas_punto_desconocido, FinalSolution1)
  error2 = calcular_error(coordenadas_punto_desconocido, FinalSolution2)

  if error1 <= error2:
    error = error1
    return FinalSolution1
  else:
    error = error2
    return FinalSolution2

## Calculo del error cometido
def calcular_error(coordenadas_conocidas, coordenadas_calculadas):

    coordenadas_conocidas_np = np.array(coordenadas_conocidas)
    coordenadas_calculadas_np = np.array(coordenadas_calculadas)
    error = np.linalg.norm(coordenadas_conocidas_np - coordenadas_calculadas_np)
    errorRounded = np.round(error,4)
    return errorRounded

def calcular_error_entre_coordenadas(coordenadas_conocidas, coordenadas_calculadas):
    error_entre_coordenadas = []

    if len(coordenadas_conocidas) == numero_coordenadas:
      num_coords = numero_coordenadas
    else:
      num_coords = len(coordenadas_conocidas)
    
    for i in range(num_coords):
      diferencia = abs(coordenadas_calculadas[i]) - abs(coordenadas_conocidas[i])
      error_entre_coordenadas.append(np.round(diferencia,4))

    return error_entre_coordenadas


## DEFINICION DE FUNCIONES PARA LA EJECUCION DE LA MULTILATERALIZACION ##
## ENTRADA DE DATOS, INICIALIZACION DE LOS DATOS, EJECUCION MULTILATERALIZACION ##

## Entrada de datos para la batería de tests ##
def entradaDeDatosUser():
  global replicarPaper, numero_puntos_ancla, numero_de_iteraciones
  replicarPaper = (input("¿Quieres replicar el paper? Escribe Si para replicarlo, No en caso contrario, Sim en caso de querer simular un caso real con valores aleatorios o Real para usar datos REALES: ")).lower()
  if replicarPaper != "si":
    print()
    print("Introducir sólo números enteros")
    numero_puntos_ancla = int(input("Número de puntos ancla: "))
    numero_de_iteraciones = int(input("Número de iteraciones: "))

## Inicializacion de los datos según la prueba que se quiera hacer: Réplica del test, mismas condiciones que el paper con subespacio 4 y valores aleatorios, Simulación app
def inicializacionDeDatos():
  global nombre_muestra_real, coordenadas_punto_desconocido, coordenadas_puntos_ancla_conocidos, distancias_REALES_a_cada_punto_ancla, distancias_MEDIDAS_a_cada_punto_ancla, numero_puntos_ancla
  if replicarPaper == "si":

    coordenadas_punto_desconocido = np.array([24.335, -2.506, 1.13])
    coordenadas_puntos_ancla_conocidos = np.array([[27.297, -4.953, 1.47], [25.475, -6.124, 2.36], [22.59, 0.524, 1.2]])

    solucion1_paper = np.array([24.3506, -2.4811, 1.6667])
    solucion2_paper = np.array([24.3123, -2.5205, 1.5365])

    distancias_MEDIDAS_a_cada_punto_ancla = np.array([3.851, 3.875, 3.514])
    distancias_REALES_a_cada_punto_ancla = np.array([3.857, 3.988, 3.497])

    error_al_MEDIR_distancias = np.array([0.6, -11.3, -1.7])
    print("#### RÉPLICA PAPER ###")
    print("#### Método de multilateralización en un espacio de 3 dimensiones, con 3 puntos Ancla, medidas las distancias al punto desconocido ###")

  else :

    infoMessage = f"#### Método de multilateralización en un espacio de {numero_coordenadas} dimensiones, con {numero_puntos_ancla} puntos ANCLA, "

    if replicarPaper == "sim":

      print(infoMessage, "simulando una entrada de datos desde la app, es decir, 2 coordenadas del punto desconocido ###")

      # coordenadas_punto_desconocido = np.random.uniform(low=-4, high=4, size=numero_coordenadas)
      coordenadas_punto_desconocido = np.random.multivariate_normal(mean, covDependency, 1).flatten()
      coordenadas_puntos_ancla_conocidos = generar_puntos_en_R2(numero_puntos_ancla, coordenadas_punto_desconocido, 1, 1)
      # coordenadas_puntos_ancla_conocidos = np.random.multivariate_normal(mean, covDependency, numero_puntos_ancla)
      print()
      print("Coordenadas conocidas del punto desconocido generadas de forma aleatoria con dependencia entre ambas: ", coordenadas_punto_desconocido)
      print("Puntos ancla generados de forma aleatoria: ", coordenadas_puntos_ancla_conocidos)

      calcular_distancias_subespacio_2D(coordenadas_puntos_ancla_conocidos, coordenadas_punto_desconocido)

    elif replicarPaper == "no":

      print(infoMessage, "medidas las distancias al punto desconocido ###")

      coordenadas_punto_desconocido = np.random.uniform(low=-10, high=10, size=numero_coordenadas)
      coordenadas_puntos_ancla_conocidos = generadorPuntosRandomConUnaDistanciaDeterminada(coordenadas_punto_desconocido, distanciaPuntoDesconocido, numero_puntos_ancla)

      print()
      print("Coordenadas del punto desconocido:", coordenadas_punto_desconocido)
      print("Coordenadas de los puntos ancla conocidos:", coordenadas_puntos_ancla_conocidos)
      print()

      calculateDistances()
      distancias_MEDIDAS_a_cada_punto_ancla = distancias_REALES_a_cada_punto_ancla

    elif replicarPaper == "real":

      # Sube el archivo CSV
      file_path = '/content/valoresReales.csv'
      df = pd.read_csv(file_path)

      # Mostrar la primera fila (índice 0)
      first_row = df.iloc[0]

      # La primera columna es el título y las siguientes 4 son las coordenadas
      nombre_muestra_real = first_row[0]

      print(infoMessage, "con DATOS REALES, simulando una entrada de datos desde la app, es decir, 2 coordenadas del punto desconocido ###")

      coordenadas_punto_desconocido = np.array([first_row[1:5].astype(float).tolist()])
      coordenadas_puntos_ancla_conocidos = np.random.multivariate_normal(mean, covHighDependency, numero_puntos_ancla)
      print()
      print("Coordenadas con valores reales: ", coordenadas_punto_desconocido)
      print("Puntos ancla generados de forma aleatoria: ", coordenadas_puntos_ancla_conocidos)

      calcular_distancias_subespacio_2D(coordenadas_puntos_ancla_conocidos, coordenadas_punto_desconocido)

## Método de multilateralizacion mediante el metodo algebraico del paper
def startMultilateralization():
  global FinalSolution, error, A, b, refPoints
  refPoints = coordenadas_puntos_ancla_conocidos
  distancias = distancias_MEDIDAS_a_cada_punto_ancla

  print("Distancias puntos Ancla con el punto desconocido: ", distancias)

  I = np.identity(len(refPoints[0])) ## Matriz Identidad

  generateAEcs()
  generateBEcs(distancias)
  A_pseudo_inv = np.linalg.pinv(A)
  x_p = A_pseudo_inv.dot(b)

  U, S, VT = np.linalg.svd(A)
  x_h = VT[-1]

  secondGradeVariables(x_p, x_h)
  insideSqrt = (B_segundoGrado**2)-(4*A_segundoGrado*C_segundoGrado)

  ecSecondGradeSolution = solve_quadratic(A_segundoGrado, B_segundoGrado, insideSqrt)

  if len(ecSecondGradeSolution) == 1:
    sol = x_p + ecSecondGradeSolution*x_h
    sol = sol[1:]
    FinalSolution = sol.dot(I)
    FinalSolutionRounded = np.round(FinalSolution, 4)
    error = calcular_error(coordenadas_punto_desconocido, FinalSolutionRounded)
    errorCoordenadas = calcular_error_entre_coordenadas(coordenadas_punto_desconocido, FinalSolutionRounded)
  else:
    sol1 = x_p + ecSecondGradeSolution[0]*x_h
    sol2 = x_p + ecSecondGradeSolution[1]*x_h
    sol1 = sol1[1:]
    sol2 = sol2[1:]
    FinalSolution1 = sol1.dot(I)
    FinalSolution2 = sol2.dot(I)
    FinalSolution1Rounded = np.round(FinalSolution1, 4)
    FinalSolution2Rounded = np.round(FinalSolution2, 4)
    print()
    print("POSIBLE SOLUCION FINAL 1 =", FinalSolution1Rounded)
    print("POSIBLE SOLUCION FINAL 2 =", FinalSolution2Rounded)
    FinalSolution = findSolution(FinalSolution1, FinalSolution2)
    FinalSolutionRounded = np.round(FinalSolution, 4)
    errorCoordenadas = calcular_error_entre_coordenadas(coordenadas_punto_desconocido, FinalSolutionRounded)

  acumulatedDistanceError.append(error) ## Una vez determinada la distancia de error cometida entre la posición del punto calculado y la posición del punto real, se guarda para calcular el MAE
  print()
  print("Las coordenadas del PUNTO DESCONOCIDO ERAN: ", coordenadas_punto_desconocido)
  print("Las coordenadas del PUNTO DESCONOCIDO CALCULADAS = ", FinalSolutionRounded)
  print("El error cometido entre coordenadas es: ", errorCoordenadas)
  print("El error cometido en la distancia es: ", error)

### PLOTS ###

def generatePlotsCoordenadasCartesianas():
    global radioCircunferenciaDistancias
    combinations = [
        ((coordenadas_punto_desconocido[0], coordenadas_punto_desconocido[1]), [(point[0], point[1]) for point in coordenadas_puntos_ancla_conocidos], 'X-Axis', 'Y-Axis'),
        ((coordenadas_punto_desconocido[2], coordenadas_punto_desconocido[3]), [(point[2], point[3]) for point in coordenadas_puntos_ancla_conocidos], 'Z-Axis', 'W-Axis'),
        ((coordenadas_punto_desconocido[0], coordenadas_punto_desconocido[3]), [(point[0], point[3]) for point in coordenadas_puntos_ancla_conocidos], 'X-Axis', 'W-Axis'),
        ((coordenadas_punto_desconocido[1], coordenadas_punto_desconocido[2]), [(point[1], point[2]) for point in coordenadas_puntos_ancla_conocidos], 'Y-Axis', 'Z-Axis')
    ]

    fig, axs = plt.subplots(2, 2, figsize=(16, 16))
    axs = axs.flatten()

    for ax, (base_coord, points_coords, xlabel, ylabel) in zip(axs, combinations):
        base_x, base_y = base_coord
        points_x = [point[0] for point in points_coords]
        points_y = [point[1] for point in points_coords]

        ax.scatter(points_x, points_y, c='blue', label='Puntos ancla')
        ax.scatter([base_x], [base_y], c='red', label='Punto desconocido')
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.set_title(f'Puntos ancla alrededor del punto desconocido ({xlabel}-{ylabel})')

        upperCircle = plt.Circle((base_x, base_y), radioCircunferenciaDistancias, color='green', fill=False, linestyle='--', label=f'Circunferencia distancia {radioCircunferenciaDistancias}')
        downCircle = plt.Circle((base_x, base_y), radioCircunferenciaDistancias, color='green', fill=False, linestyle='--', label=f'Circunferencia distancia {radioCircunferenciaDistancias}')
        circleUnity = plt.Circle((base_x, base_y), 1, color='orange', fill=False, linestyle='--', label='Circunferencia Unidad')

        ax.add_patch(upperCircle)
        ax.add_patch(downCircle)
        ax.add_patch(circleUnity)
        ax.set_aspect('equal', adjustable='box')
        ax.legend(loc='upper right')
        ax.grid(True)


    plt.tight_layout()
    plt.show()

## Funciones principales que llevan a cabo el flujo para la multilateralizacion ##
entradaDeDatosUser()

for i in range(0,numero_de_iteraciones):
  print()
  print()
  print(f"Ejecución {i+1} de {numero_de_iteraciones}")
  print()
  inicializacionDeDatos()
  start_time = time.time()
  startMultilateralization()
  if replicarPaper != "si":
    generatePlotsCoordenadasCartesianas()
  end_time = time.time()
  performance_multilateralizacion.append(end_time - start_time)

# files.download('output_log.txt')

mean_distance_error = round(np.mean(acumulatedDistanceError), 2)
mean_performance = round(np.mean(performance_multilateralizacion), 4)

########################


########################

### EXCEL con los datos de las iteraciones ###

def generar_excel(nombres_variables, maeDistances, meanPerformances, ruta_archivo):
      workbook = Workbook()

      worksheet = workbook.active
      worksheet.title = "Test Multilateralizacion"

      for col, nombre_variable in enumerate(nombres_variables):
          worksheet.cell(row=1, column=col + 2, value=nombre_variable)

      column_performance = len(nombres_variables)
      worksheet.cell(row=2, column=column_performance, value=meanPerformances)

      worksheet.cell(row=2, column=len(nombres_variables) - 1, value=maeDistances)
      worksheet.cell(row=2, column=len(nombres_variables) + 1, value=maeDistances*100)

      workbook.save(ruta_archivo)
      # Descarga el archivo Excel generado
      files.download(ruta_archivo)

if replicarPaper != "si":
  ruta_archivo = f"Test_{numero_puntos_ancla}Anchors&{numero_de_iteraciones}Loops_{replicarPaper.upper()}.xlsx"
  nombres_variables = ["MAE Distances", "Performance (s)", "Error %"]

  generar_excel(nombres_variables, mean_distance_error, mean_performance, ruta_archivo)
