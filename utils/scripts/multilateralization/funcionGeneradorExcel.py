from openpyxl import Workbook

numero_puntos_ancla = 4
numero_de_iteraciones = 4
replicarPaper = "??"
ruta_archivo = f"Test_{numero_puntos_ancla}Anchors&{numero_de_iteraciones}Loops_{replicarPaper.upper()}.xlsx"
nombres_variables = ["MAE Distances", "Performance (s)", "Error %"]
toInsert = []

def generar_excel(nombres_variables, maeDistances, meanPerformances, ruta_archivo):
      workbook = Workbook()

      worksheet = workbook.active
      worksheet.title = "Test Multilateralizacion"

      for col, nombre_variable in enumerate(nombres_variables):
          worksheet.cell(row=1, column=col + 2, value=nombre_variable)

      column_performance = len(nombres_variables)
      worksheet.cell(row=2, column=column_performance, value=meanPerformances)

      worksheet.cell(row=2, column=len(nombres_variables) - 1, value=maeDistances)
      worksheet.cell(row=2, column=len(nombres_variables) + 1, value=maeDistances*100)

      workbook.save(ruta_archivo)
      # Descarga el archivo Excel generado
      files.download(ruta_archivo)


generar_excel(nombres_variables, mean_distance_error, mean_performance, ruta_archivo)